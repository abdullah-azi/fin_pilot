import csv
import base64
import hashlib
import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from uuid import UUID

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.import_batch import ImportBatch
from app.models.enums import TransactionFrequency, TransactionType
from app.models.transaction import Transaction
from app.schemas.imports import (
    ImportBatchHistoryItem,
    ImportBatchHistoryResponse,
    CSVImportConfirmRequest,
    CSVImportPreviewResponse,
    CSVImportPreviewRow,
    CSVImportSkippedRow,
)
from app.services.category_matching import build_category_lookup, infer_category_match
from app.services.categories import list_categories
from app.services.transactions import MONEY_QUANTUM, _ensure_category_access

DATE_HEADERS = (
    "date",
    "transaction_date",
    "posted_date",
    "posting_date",
    "created_at",
    "timestamp",
    "time",
)
TITLE_HEADERS = (
    "description",
    "details",
    "merchant",
    "title",
    "transaction",
    "narration",
    "remarks",
    "reference",
    "name",
)
AMOUNT_HEADERS = ("amount", "transaction_amount", "value", "amount_pkr")
DEBIT_HEADERS = ("debit", "withdrawal", "money_out", "debit_amount", "outflow")
CREDIT_HEADERS = ("credit", "deposit", "money_in", "credit_amount", "inflow")
TYPE_HEADERS = ("type", "transaction_type", "direction", "dr_cr", "drcr")
NOTE_HEADERS = ("note", "notes", "memo", "reference", "remarks")
BALANCE_HEADERS = ("balance", "closing_balance", "available_balance")


@dataclass(slots=True)
class ParsedImportRow:
    row_index: int
    transaction_date: date
    title: str
    amount: Decimal
    balance: Decimal | None
    transaction_type: TransactionType
    note: str | None
    category_id: UUID | None
    category_name: str | None
    fingerprint: str
    raw_preview: dict[str, str]


@dataclass(slots=True)
class SkippedImportRow:
    row_index: int
    reason: str
    raw_preview: dict[str, str]


@dataclass(slots=True)
class PreviewCSVResult:
    source_name: str | None
    detected_columns: list[str]
    rows: list[ParsedImportRow]
    skipped_rows: list[SkippedImportRow]


@dataclass(slots=True)
class ConfirmCSVResult:
    source_name: str | None
    imported_count: int
    skipped_duplicate_count: int
    skipped_duplicates: list[int]
    imported_transaction_ids: list[UUID]


def preview_csv_import(
    db: Session,
    *,
    user_id: UUID,
    file_name: str | None,
    file_bytes: bytes,
) -> PreviewCSVResult:
    decoded = _decode_csv_bytes(file_bytes)
    dialect = _sniff_dialect(decoded)
    detected_columns, data_rows = _extract_header_and_data_rows(decoded, dialect)
    return _build_preview_result(
        db,
        user_id=user_id,
        file_name=file_name,
        detected_columns=detected_columns,
        data_rows=data_rows,
    )


def preview_xlsx_import(
    db: Session,
    *,
    user_id: UUID,
    file_name: str | None,
    content_base64: str,
) -> PreviewCSVResult:
    try:
        file_bytes = base64.b64decode(content_base64, validate=True)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file content is not valid base64.",
        ) from exc

    try:
        workbook = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded Excel file could not be read.",
        ) from exc

    worksheet = workbook.active
    spreadsheet_rows = [
        [_stringify_spreadsheet_cell(cell) for cell in row]
        for row in worksheet.iter_rows(values_only=True)
    ]

    detected_columns, data_rows = _extract_header_and_data_rows_from_records(spreadsheet_rows)
    return _build_preview_result(
        db,
        user_id=user_id,
        file_name=file_name,
        detected_columns=detected_columns,
        data_rows=data_rows,
    )


def build_preview_response(result: PreviewCSVResult) -> CSVImportPreviewResponse:
    return CSVImportPreviewResponse(
        source_name=result.source_name,
        detected_columns=result.detected_columns,
        parsed_count=len(result.rows),
        skipped_count=len(result.skipped_rows),
        rows=[
            CSVImportPreviewRow(
                row_index=row.row_index,
                transaction_date=row.transaction_date,
                title=row.title,
                amount=row.amount,
                balance=row.balance,
                type=row.transaction_type,
                note=row.note,
                category_id=row.category_id,
                category_name=row.category_name,
                fingerprint=row.fingerprint,
                raw_preview=row.raw_preview,
            )
            for row in result.rows
        ],
        skipped_rows=[
            CSVImportSkippedRow(
                row_index=row.row_index,
                reason=row.reason,
                raw_preview=row.raw_preview,
            )
            for row in result.skipped_rows
        ],
    )


def confirm_csv_import(
    db: Session,
    *,
    user_id: UUID,
    payload: CSVImportConfirmRequest,
) -> ConfirmCSVResult:
    if not payload.rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="At least one parsed row is required for import.",
        )

    for row in payload.rows:
        if row.category_id is not None:
            _ensure_category_access(db, user_id, row.category_id)

    min_date = min(row.transaction_date for row in payload.rows)
    max_date = max(row.transaction_date for row in payload.rows)
    original_parsed_count = payload.original_parsed_count if payload.original_parsed_count is not None else len(payload.rows)
    ignored_count = max(0, original_parsed_count - len(payload.rows))

    existing_rows = db.execute(
        select(Transaction).where(
            Transaction.user_id == user_id,
            Transaction.transaction_date >= min_date,
            Transaction.transaction_date <= max_date,
        )
    ).scalars().all()

    existing_keys = {
        _build_duplicate_key(
            transaction.transaction_date,
            transaction.type,
            transaction.amount,
            transaction.title,
            _extract_balance_from_note(transaction.note),
        )
        for transaction in existing_rows
    }
    existing_fingerprints = {
        _row_fingerprint(
            transaction.transaction_date,
            transaction.type,
            transaction.amount,
            transaction.title,
            _extract_balance_from_note(transaction.note),
        )
        for transaction in existing_rows
    }

    seen_payload_keys: set[tuple[date, str, str, str, str | None]] = set()
    seen_payload_fingerprints: set[str] = set()
    imported_transaction_ids: list[UUID] = []
    skipped_duplicates: list[int] = []

    for row in payload.rows:
        if row.fingerprint in existing_fingerprints or row.fingerprint in seen_payload_fingerprints:
            skipped_duplicates.append(row.row_index)
            continue

        duplicate_key = _build_duplicate_key(
            row.transaction_date,
            row.type,
            row.amount.quantize(MONEY_QUANTUM),
            row.title,
            row.balance.quantize(MONEY_QUANTUM) if row.balance is not None else None,
        )
        if duplicate_key in existing_keys or duplicate_key in seen_payload_keys:
            skipped_duplicates.append(row.row_index)
            continue

        transaction = Transaction(
            user_id=user_id,
            type=row.type,
            amount=row.amount.quantize(MONEY_QUANTUM),
            income_frequency=TransactionFrequency.ONCE if row.type == TransactionType.INCOME else None,
            hours_per_day=None,
            days_per_week=None,
            category_id=row.category_id,
            title=row.title.strip(),
            note=row.note,
            transaction_date=row.transaction_date,
        )
        db.add(transaction)
        db.flush()

        imported_transaction_ids.append(transaction.id)
        existing_keys.add(duplicate_key)
        seen_payload_keys.add(duplicate_key)
        existing_fingerprints.add(row.fingerprint)
        seen_payload_fingerprints.add(row.fingerprint)

    import_batch = ImportBatch(
        user_id=user_id,
        source_name=payload.source_name,
        original_parsed_count=original_parsed_count,
        requested_count=len(payload.rows),
        imported_count=len(imported_transaction_ids),
        ignored_count=ignored_count,
        skipped_duplicate_count=len(skipped_duplicates),
        transaction_date_from=min_date,
        transaction_date_to=max_date,
    )
    db.add(import_batch)

    db.commit()

    return ConfirmCSVResult(
        source_name=payload.source_name,
        imported_count=len(imported_transaction_ids),
        skipped_duplicate_count=len(skipped_duplicates),
        skipped_duplicates=skipped_duplicates,
        imported_transaction_ids=imported_transaction_ids,
    )


def list_import_batches(
    db: Session,
    *,
    user_id: UUID,
    limit: int = 20,
) -> ImportBatchHistoryResponse:
    rows = db.execute(
        select(ImportBatch)
        .where(ImportBatch.user_id == user_id)
        .order_by(ImportBatch.created_at.desc())
        .limit(limit)
    ).scalars().all()

    return ImportBatchHistoryResponse(
        items=[
            ImportBatchHistoryItem(
                id=row.id,
                source_name=row.source_name,
                original_parsed_count=row.original_parsed_count,
                requested_count=row.requested_count,
                imported_count=row.imported_count,
                ignored_count=row.ignored_count,
                skipped_duplicate_count=row.skipped_duplicate_count,
                transaction_date_from=row.transaction_date_from,
                transaction_date_to=row.transaction_date_to,
                created_at=row.created_at.isoformat(),
                updated_at=row.updated_at.isoformat(),
            )
            for row in rows
        ]
    )


def _parse_csv_row(
    *,
    row_index: int,
    raw_row: dict[str, str | None],
    normalized_header_map: dict[str, str],
    category_lookup,
) -> ParsedImportRow:
    transaction_date = _extract_date(raw_row, normalized_header_map)
    title = _extract_title(raw_row, normalized_header_map)
    amount, transaction_type = _extract_amount_and_type(raw_row, normalized_header_map)
    balance = _extract_balance(raw_row, normalized_header_map)
    note = _extract_note(raw_row, normalized_header_map)
    category_id, category_name = infer_category_match(
        title=title,
        note=note,
        transaction_type=transaction_type,
        category_lookup=category_lookup,
    )

    raw_preview = {
        key.strip(): (value or "").strip()
        for key, value in raw_row.items()
        if key and ((value or "").strip())
    }

    return ParsedImportRow(
        row_index=row_index,
        transaction_date=transaction_date,
        title=title,
        amount=amount,
        balance=balance,
        transaction_type=transaction_type,
        note=note,
        category_id=category_id,
        category_name=category_name,
        fingerprint=_row_fingerprint(transaction_date, transaction_type, amount, title, balance),
        raw_preview=raw_preview,
    )


def _decode_csv_bytes(file_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="CSV file encoding is not supported.",
    )


def _extract_header_and_data_rows(
    content: str,
    dialect: csv.Dialect,
) -> tuple[list[str], list[tuple[int, dict[str, str]]]]:
    reader = csv.reader(io.StringIO(content), dialect=dialect)
    return _extract_header_and_data_rows_from_records(reader)


def _extract_header_and_data_rows_from_records(
    records,
) -> tuple[list[str], list[tuple[int, dict[str, str]]]]:
    header_row: list[str] | None = None
    data_rows: list[tuple[int, dict[str, str]]] = []

    for record_index, row in enumerate(records, start=1):
        cleaned_row = [cell.strip() for cell in row]

        if header_row is None:
            if _looks_like_header_row(cleaned_row):
                header_row = cleaned_row
            continue

        row_mapping = {
            header_row[index]: row[index] if index < len(row) else ""
            for index in range(len(header_row))
            if header_row[index]
        }
        data_rows.append((record_index, row_mapping))

    if header_row is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV file has no usable header row.",
        )

    return [field for field in header_row if field], data_rows


def _sniff_dialect(content: str) -> csv.Dialect:
    sample = content[:2048]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        return csv.get_dialect("excel")


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")


def _stringify_spreadsheet_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _build_preview_result(
    db: Session,
    *,
    user_id: UUID,
    file_name: str | None,
    detected_columns: list[str],
    data_rows: list[tuple[int, dict[str, str]]],
) -> PreviewCSVResult:
    normalized_header_map = {
        _normalize_header(field_name): field_name
        for field_name in detected_columns
        if field_name is not None
    }

    if not _has_any_header(normalized_header_map, DATE_HEADERS):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Statement must include a date column.",
        )

    if not (
        _has_any_header(normalized_header_map, AMOUNT_HEADERS)
        or _has_any_header(normalized_header_map, DEBIT_HEADERS)
        or _has_any_header(normalized_header_map, CREDIT_HEADERS)
    ):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Statement must include an amount column or debit/credit columns.",
        )

    categories = list_categories(db, user_id, include_hidden=True)
    category_lookup = build_category_lookup(categories)

    parsed_rows: list[ParsedImportRow] = []
    skipped_rows: list[SkippedImportRow] = []

    for row_index, raw_row in data_rows:
        compact_preview = {
            key.strip(): (value or "").strip()
            for key, value in raw_row.items()
            if key and ((value or "").strip())
        }

        if not compact_preview:
            continue

        try:
            parsed_rows.append(
                _parse_csv_row(
                    row_index=row_index,
                    raw_row=raw_row,
                    normalized_header_map=normalized_header_map,
                    category_lookup=category_lookup,
                )
            )
        except ValueError as exc:
            skipped_rows.append(
                SkippedImportRow(
                    row_index=row_index,
                    reason=str(exc),
                    raw_preview=compact_preview,
                )
            )

    return PreviewCSVResult(
        source_name=file_name,
        detected_columns=detected_columns,
        rows=parsed_rows,
        skipped_rows=skipped_rows,
    )


def _looks_like_header_row(row: list[str]) -> bool:
    normalized_cells = {_normalize_header(cell) for cell in row if cell.strip()}
    has_date = any(candidate in normalized_cells for candidate in DATE_HEADERS)
    has_title = any(candidate in normalized_cells for candidate in TITLE_HEADERS)
    has_amount = any(candidate in normalized_cells for candidate in AMOUNT_HEADERS + DEBIT_HEADERS + CREDIT_HEADERS)
    return has_date and has_title and has_amount


def _has_any_header(normalized_header_map: dict[str, str], candidates: tuple[str, ...]) -> bool:
    return any(candidate in normalized_header_map for candidate in candidates)


def _find_value(
    raw_row: dict[str, str | None],
    normalized_header_map: dict[str, str],
    candidates: tuple[str, ...],
) -> str | None:
    for candidate in candidates:
        header = normalized_header_map.get(candidate)
        if header is None:
            continue

        value = raw_row.get(header)
        if value is not None and value.strip():
            return value.strip()

    return None


def _extract_date(raw_row: dict[str, str | None], normalized_header_map: dict[str, str]) -> date:
    value = _find_value(raw_row, normalized_header_map, DATE_HEADERS)
    if not value:
        raise ValueError("Missing transaction date.")

    parsed = _parse_date(value)
    if parsed is None:
        raise ValueError(f"Could not parse date: {value}")

    return parsed


def _extract_title(raw_row: dict[str, str | None], normalized_header_map: dict[str, str]) -> str:
    value = _find_value(raw_row, normalized_header_map, TITLE_HEADERS)
    if not value:
        raise ValueError("Missing transaction description/title.")

    return value[:120]


def _extract_amount_and_type(
    raw_row: dict[str, str | None],
    normalized_header_map: dict[str, str],
) -> tuple[Decimal, TransactionType]:
    explicit_type_value = _find_value(raw_row, normalized_header_map, TYPE_HEADERS)
    title_value = _find_value(raw_row, normalized_header_map, TITLE_HEADERS)
    explicit_type = (
        _parse_type_value(explicit_type_value, title_hint=title_value)
        if explicit_type_value
        else _infer_transaction_type_from_text(title_value)
    )

    amount_value = _find_value(raw_row, normalized_header_map, AMOUNT_HEADERS)
    if amount_value:
        amount = _parse_decimal(amount_value)
        if amount is None:
            raise ValueError(f"Could not parse amount: {amount_value}")

        if explicit_type is not None:
            return abs(amount).quantize(MONEY_QUANTUM), explicit_type

        if amount < 0:
            return abs(amount).quantize(MONEY_QUANTUM), TransactionType.EXPENSE
        if amount > 0:
            return amount.quantize(MONEY_QUANTUM), TransactionType.INCOME

    debit_value = _find_value(raw_row, normalized_header_map, DEBIT_HEADERS)
    credit_value = _find_value(raw_row, normalized_header_map, CREDIT_HEADERS)

    debit_amount = _parse_decimal(debit_value) if debit_value else None
    credit_amount = _parse_decimal(credit_value) if credit_value else None

    if debit_amount and debit_amount > 0:
        return debit_amount.quantize(MONEY_QUANTUM), TransactionType.EXPENSE
    if credit_amount and credit_amount > 0:
        return credit_amount.quantize(MONEY_QUANTUM), TransactionType.INCOME

    raise ValueError(
        f"Could not determine whether this row is income or expense. "
        f"type={explicit_type_value or 'missing'} amount={amount_value or 'missing'} "
        f"debit={debit_value or 'missing'} credit={credit_value or 'missing'}"
    )


def _extract_note(raw_row: dict[str, str | None], normalized_header_map: dict[str, str]) -> str | None:
    parts: list[str] = []
    note_value = _find_value(raw_row, normalized_header_map, NOTE_HEADERS)
    balance_value = _find_value(raw_row, normalized_header_map, BALANCE_HEADERS)

    if note_value:
        parts.append(note_value)
    if balance_value:
        parts.append(f"Balance: {balance_value}")

    return " | ".join(parts)[:255] if parts else None


def _extract_balance(raw_row: dict[str, str | None], normalized_header_map: dict[str, str]) -> Decimal | None:
    balance_value = _find_value(raw_row, normalized_header_map, BALANCE_HEADERS)
    if not balance_value:
        return None

    balance = _parse_decimal(balance_value)
    if balance is None:
        return None
    return balance.quantize(MONEY_QUANTUM)


def _parse_date(value: str) -> date | None:
    normalized = value.strip()

    iso_candidate = normalized[:10]
    try:
        return date.fromisoformat(iso_candidate)
    except ValueError:
        pass

    datetime_formats = (
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
        "%d %b %Y",
        "%d %B %Y",
        "%d/%m/%Y %H:%M",
        "%m/%d/%Y %H:%M",
        "%d %b %Y %I:%M %p",
        "%d %B %Y %I:%M %p",
        "%d %b %Y %H:%M",
        "%d %B %Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M:%S",
        "%d %b %Y %I:%M:%S %p",
        "%d %B %Y %I:%M:%S %p",
        "%d %b %Y %H:%M:%S",
        "%d %B %Y %H:%M:%S",
    )
    for fmt in datetime_formats:
        try:
            return datetime.strptime(normalized, fmt).date()
        except ValueError:
            continue

    return None


def _parse_decimal(value: str | None) -> Decimal | None:
    if value is None:
        return None

    cleaned = value.strip()
    if not cleaned:
        return None

    negative = False
    if cleaned.startswith("(") and cleaned.endswith(")"):
        negative = True
        cleaned = cleaned[1:-1]

    cleaned = cleaned.replace(",", "")
    cleaned = re.sub(r"(?i)(pkr|rs\.?|usd|eur|qar)", "", cleaned).strip()
    cleaned = re.sub(r"[^0-9.\-]", "", cleaned)
    if not cleaned:
        return None

    try:
        amount = Decimal(cleaned)
    except InvalidOperation:
        return None

    if negative:
        amount = -amount

    return amount


def _parse_type_value(value: str, *, title_hint: str | None = None) -> TransactionType | None:
    normalized = _normalize_header(value)
    if normalized in {"expense", "debit", "dr", "withdrawal", "money_out"}:
        return TransactionType.EXPENSE
    if normalized in {"income", "credit", "cr", "deposit", "money_in"}:
        return TransactionType.INCOME
    if normalized in {"raast_in", "ibft_in", "incoming_transfer", "funds_received"}:
        return TransactionType.INCOME
    if normalized in {"raast_out", "cash_withdrawal", "cash_out", "pos", "mobile_top_up"}:
        return TransactionType.EXPENSE
    if normalized == "peer_to_peer":
        return _infer_transaction_type_from_text(title_hint)
    return _infer_transaction_type_from_text(title_hint or value)


def _infer_transaction_type_from_text(value: str | None) -> TransactionType | None:
    if not value:
        return None

    normalized = value.strip().lower()
    income_markers = (
        "incoming",
        "received",
        "receive",
        "salary",
        "deposit",
        "fund transfer from",
    )
    expense_markers = (
        "outgoing",
        "sent",
        "send",
        "paid to",
        "cash withdrawn",
        "top-up purchased",
        "withdrawal",
    )

    if any(marker in normalized for marker in income_markers):
        return TransactionType.INCOME
    if any(marker in normalized for marker in expense_markers):
        return TransactionType.EXPENSE
    return None


def _row_fingerprint(
    transaction_date: date,
    transaction_type: TransactionType,
    amount: Decimal,
    title: str,
    balance: Decimal | None = None,
) -> str:
    balance_part = f"|{balance.quantize(MONEY_QUANTUM)}" if balance is not None else ""
    raw = (
        f"{transaction_date.isoformat()}|{transaction_type.value}|"
        f"{amount.quantize(MONEY_QUANTUM)}|{title.strip().lower()}{balance_part}"
    )
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _build_duplicate_key(
    transaction_date: date,
    transaction_type: TransactionType,
    amount: Decimal,
    title: str,
    balance: Decimal | None = None,
) -> tuple[date, str, str, str, str | None]:
    return (
        transaction_date,
        transaction_type.value,
        f"{amount.quantize(MONEY_QUANTUM)}",
        title.strip().lower(),
        f"{balance.quantize(MONEY_QUANTUM)}" if balance is not None else None,
    )


def _extract_balance_from_note(note: str | None) -> Decimal | None:
    if not note:
        return None

    match = re.search(r"balance:\s*([^\|]+)$", note, flags=re.IGNORECASE)
    if not match:
        return None

    return _parse_decimal(match.group(1))
